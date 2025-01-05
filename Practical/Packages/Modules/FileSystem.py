import os
from openpyxl import Workbook, load_workbook
import time
import re
import csv



def FileExists(Assumption:str ,GroundTruth:str):
    if os.path.exists(Assumption):
        os.remove(Assumption)
    if os.path.exists(GroundTruth):
        os.remove(GroundTruth)



def Read_GroundTruth(reference_file:str):

    extracted_data = []

    # Open the file and read it
    with open(reference_file, 'r') as file:
        reader = csv.reader(file)  # Create a CSV reader
        headers = next(reader)  # Read the header row (optional)

        # Iterate through each row in the CSV
        for i, row in enumerate(reader):
            if i >= 30:  # Stop after 40 rows
                break
            # Fetch columns 3, 4, 5, 6, 7 (indexes 2, 3, 4, 5, 6)
            extracted_row = row[7:13]
            # Access each element separately and convert each to int, if needed.
            Converted_Data = []
            for row in extracted_row:
                Converted_Data.append(int(row))
            # Slicing to get columns 3 to 7
            extracted_data.append(Converted_Data)
    return extracted_data

def WriteGroundTruth(GroundTruth_file ,extracted_data):
    if os.path.exists(GroundTruth_file):
        # Load the existing workbook
        workbook = load_workbook(GroundTruth_file)
        sheet = workbook.active
        print("File exists. Opened in append mode.")
    else:
        # Create a new workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"  # Optionally, rename the sheet
        print("File created.")

    # Example of adding data to the workbook
    sheet.append(["Prolongation", "Block", "SoundRep", "WordRep", "DifficultToUnderstand", "Interjection"])
    for item in extracted_data:
        sheet.append(item)

    # Save the file
    workbook.save(GroundTruth_file)
    print(f"Data written to {GroundTruth_file}")