import os
from dotenv import load_dotenv
import asyncio
import time
from Modules.Whisper import call_whisper,call_whisper2
from openpyxl import Workbook, load_workbook
import re
import ssl
from dotenv import load_dotenv
from Practical.Packages.Modules.Accuracy import find_Accuracy
from Practical.Packages.Modules.FileSystem import FileExists, Read_GroundTruth, WriteGroundTruth
from Practical.Packages.Modules.File_List import file_List
from Practical.Packages.Modules.LLM import LLM_search
import warnings

warnings.filterwarnings("ignore", message="FP16 is not supported on CPU; using FP32 instead")
load_dotenv()


Assumption_file = "Assumption/Assumption.xlsx"
GroundTruth_file = "GroundTruth/GroundTruth.xlsx"
Reference_file = "Data/SEP-28k_labels.csv"
path = 'clom'
File_List_Path = "File_List/File_List.json"
Transcription_File_Path = "Transcrition/Transcrition.json"





FileExists(Assumption_file,GroundTruth_file,File_List_Path,Transcription_File_Path)
ssl._create_default_https_context = ssl._create_unverified_context
file_dict = file_List(path,File_List_Path)


ModelInfor = asyncio.run(call_whisper2(File_List_Path,Transcription_File_Path))



def WriteAssumptionFile(Assumption_file:str , ModelInfor:dict):
    if os.path.exists(Assumption_file):
        workbook = load_workbook(Assumption_file)
        sheet = workbook.active
        print("File exists. Opened in append mode.")
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        print("File created.")

    sheet.append(["Prolongation", "Block", "SoundRep", "WordRep", "DifficultToUnderstand", "Interjection"])

    counter = 0
    for itemNum, itemName in ModelInfor.items():
        counter = counter + 1
        if counter % 9 == 0:
            time.sleep(60)

        print(itemName)
        bokl = LLM_search(itemName)
        data = list(map(int, re.findall(r'\b\d+\b', bokl)))
        sheet.append(data)
        print(f'{itemNum} : {bokl}')

    workbook.save(Assumption_file)

#WriteAssumptionFile(Assumption_file,ModelInfor)
#extracted_data = Read_GroundTruth(Reference_file)
#WriteGroundTruth(GroundTruth_file,extracted_data)
#find_Accuracy(Assumption_file,GroundTruth_file)
